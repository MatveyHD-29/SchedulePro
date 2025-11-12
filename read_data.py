import openpyxl
from openpyxl import load_workbook


def load_sheet(file_name, sheet_name):
    wb = load_workbook(file_name)
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        return False
    mas = []
    for row in sheet.iter_rows():
        if row[0].value != None:
            vr = []
            for cell in row:
                cell_value = cell.value
                if cell_value != None:
                    vr.append(str(cell_value))
            mas.append(vr)
        else:
            break
    return mas

def get_time_for_moving(file_name):
    mas = load_sheet(file_name, "Корпуса")
    time_for_moving = {}
    for name in mas[0]:
        if name != "-":
            for build in mas:
                if build[0] == name:
                    vr = {}
                    c = 0
                    for hour in build:
                        if hour != name and hour != "-":
                            vr[mas[0][c]] = int(hour)
                        c += 1
                    time_for_moving[name] = vr
    return time_for_moving

def get_data_from_billing(file_name):
    mas = load_sheet(file_name, "Тарификация")
    # Классы
    class_list = mas[0][2:-4]
    info_about_classes = {}
    for class_name in class_list:
        info_about_class = {}
        teachers_list_for_class = []
        class_index = mas[0].index(class_name)
        for row in mas[1:-3]:
            if row[class_index] != "-":
                teachers_list_for_class.append(row[0])
        info_about_class["Учителя"] = teachers_list_for_class
        info_about_class["Корпус"] = mas[-2][class_index]
        info_about_classes[class_name] = info_about_class
    # Учителя
    teachers_list = [i[0] for i in mas[1:-1]]
    info_about_teachers = {}
    for name in teachers_list:
        info_about_teachers_subjects = {}
        info_about_subjects = {}
        for row in mas[1:-1]:
            if row[0] == name:
                subject = row[1]
                hour_for_classes = {}
                for i in range(2, len(mas[0]) - 4):
                    if row[i] != "-":
                        hour_for_classes[list(info_about_classes.keys())[i - 2]] = row[i]
                # Исключения для "Сдвоенные уроки"
                if ";" in row[-3]:
                    st = row[-3].split("; ")
                    exceptions_double_lessons = {parametr.split(", ")[0]: int(parametr.split(", ")[1]) for parametr in st}
                elif len(row[-3]) > 1:
                    st = row[-3].split(", ")
                    exceptions_double_lessons = {st[0]: int(st[1])}
                else:
                    exceptions_double_lessons = "-"
                # Исключения для "Уроки по группам"
                if ";" in row[-1]:
                    st = row[-1].split("; ")
                    exceptions_lesson_by_groups = []
                    for parametr in st:
                        meaning = parametr.split(", ")[3] == "True"
                        vr = {}
                        vr = {"Класс": parametr.split(", ")[0], "Предмет": parametr.split(", ")[1], "Кол-во часов": parametr.split(", ")[2], "Группы чередуются": meaning}
                        exceptions_lesson_by_groups.append(vr)
                elif len(row[-1]) > 1:
                    st = row[-1].split(", ")
                    exceptions_lesson_by_groups = {"Класс": st[0], "Предмет": st[1], "Кол-во часов": st[2], "Группы чередуются": bool(st[3])}
                else:
                    exceptions_lesson_by_groups = "-"
                info_about_subjects[subject] = {"Классы": hour_for_classes, 
                                                "Сдвоенные уроки": {"Значение": True if row[-4].lower() == "да" else False, "Исключения": exceptions_double_lessons}, 
                                                "Уроки по группам": {"Значение": True if row[-2].lower() == "да" else False, "Исключения": exceptions_lesson_by_groups}}
        info_about_teachers_subjects["Предметы"] = info_about_subjects
        info_about_teachers[name] = info_about_teachers_subjects # Добавить кабинеты
    mas_teacher = load_sheet(file_name, "Учителя")
    teachers_list = [i[0] for i in mas_teacher[1:-1]]
    for row in mas_teacher[1:]:
        print(row)
        info_about_teachers[row[0]]["Кабинет"] = row[1]
        for i in range(2, 7):
            if "(" in row[i]:
                pass
            elif ";" in row[i]:
                s = row[i].split("; ")
                vr_mas = []
                for limit in s:
                    vr_mas.append(limit.split(", "))
                info_about_teachers[row[0]][mas_teacher[0][i]] = (int(vr_mas[0]), vr_mas[1])
    print(info_about_teachers["Андрюшевич М.С."])
    mas_auditorium = load_sheet(file_name, "Кабинеты")
    return info_about_classes, info_about_teachers

a = get_data_from_billing("5E22B510.xlsx")