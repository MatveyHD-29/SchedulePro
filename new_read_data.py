import openpyxl
from openpyxl import load_workbook

def load_sheet(file_name, sheet_name, method):
    wb = load_workbook(file_name)
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        return False
    mas = []
    s = sheet.iter_rows() if method == "rows" else sheet.iter_cols()
    for row in s:
        if row[0].value != None:
            vr = []
            for cell in row:
                cell_value = cell.value
                vr.append(str(cell_value))
            mas.append(vr)
        else:
            break
    return mas

def get_time_for_moving(file_name):
    mas = load_sheet(file_name, "Корпуса", "rows")
    time_for_moving = {}
    for name in mas[0]:
        if name != "-":
            for build in mas:
                if build[0] == name:
                    vr = {}
                    c = 0
                    for hour in build:
                        if hour != name and hour != "-":
                            vr[mas[0][c]] = int(hour)
                        c += 1
                    time_for_moving[name] = vr
    return time_for_moving

def load_teachers_and_classes(file_name):
    mas = load_sheet(file_name, "База данных", "cols")
    teachers_list = []
    for row in mas[0][1:]:
        if row != "None":
            teachers_list.append(row)
    classes_list = []
    for row in mas[1][1:]:
        if row != "None":
            classes_list.append(row)
    return teachers_list, classes_list

def load_info_about_teachers(file_name):
    teacher_list, classes_list = load_teachers_and_classes(file_name)
    mas_data_base = load_sheet(file_name, "Тарификация", "rows")
    info_about_teachers = {} # return
    info_about_classes = {} # return
    for name_class in classes_list:
        info_about_classes[name_class] = {}
        info_about_classes[name_class]["Предметы"] = {}
    hour_for_classes = {}
    # Добавление данных из "Тарификация" в info_about_teachers -> {"Предметы": {"Классы": {}, "Деление по группам": {}, "Урок сдвоенный": {}}}
    # Добавление данных из "Тарификация" в info_about_classes -> {"Предметы": {(Предмет): {(Учитель): (Кол-во часов)}}, "Корпус": {(Корпус)}
    for name_teacher in teacher_list:
        info_about_teachers_subjects = {}
        info_about_subjects = {}
        for row in mas_data_base[1:-1]:
            if row[0] == name_teacher:
                c = 0
                subject = row[1]
                for hour in row[2:-4]:
                    if hour != "None":
                        hour_for_classes[classes_list[c]] = hour
                        info_about_classes[classes_list[c]]["Предметы"][row[1]] = {row[0]: hour}
                    c += 1
                exception_group = row[-3]
                exception_classes = row[-1]
                info_about_subjects[subject] = {"Классы": hour_for_classes, 
                                                "Деление по группам": {"Значение": True if row[-4] == "Да" else False, "Исключения": exception_group},
                                                "Урок сдвоенный": {"Значение": True if row[-2] == "Да" else False, "Исключения": exception_classes}}
                hour_for_classes = {}
        info_about_teachers_subjects["Предметы"] = info_about_subjects
        info_about_teachers[name_teacher] = info_about_teachers_subjects
    # Добавление данных из "Тарификация" в info_about_teachers -> {"Кабинеты": {(Корпус): {(Кабинет): (Вместимость)}}
    mas_auditoriums = load_sheet(file_name, "Кабинеты", "rows")
    for row in mas_auditoriums[1:]:
        auditorium = {}
        for index_auditorium in range(1, len(row), 2):
            if row[index_auditorium] != "None" and row[index_auditorium] != "-":
                auditorium[mas_auditoriums[0][index_auditorium]] = {row[index_auditorium]: row[index_auditorium + 1]}
        info_about_teachers[row[0]]["Кабинеты"] = auditorium
    return info_about_teachers, info_about_classes

info_about_teachers, info_about_classes = load_info_about_teachers("aaa.xlsx")
print(info_about_classes["8к"])
