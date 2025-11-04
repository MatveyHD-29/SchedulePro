import openpyxl
from openpyxl import load_workbook


def load_sheet(file_name, sheet_name):
    wb = load_workbook(file_name)
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        return False
    mas = []
    for row in sheet.iter_rows():
        if row[0].value != None:
            vr = []
            for cell in row:
                cell_value = cell.value
                if cell_value != None:
                    vr.append(str(cell_value))
            mas.append(vr)
        else:
            break
    return mas

def get_time_for_moving():
    mas = load_sheet("aaa.xlsx", "Корпуса")
    time_for_moving = {}
    for name in mas[0]:
        if name != "-":
            for build in mas:
                if build[0] == name:
                    vr = {}
                    c = 0
                    for hour in build:
                        if hour != name and hour != "-":
                            vr[mas[0][c]] = int(hour)
                        c += 1
                    time_for_moving[name] = vr
    return time_for_moving

def get_data_from_billing(days):
    mas = load_sheet("aaa.xlsx", "Тарификация")
    # Классы
    info_about_classes = {mas[0][i]: mas[-1][i] for i in range(2, len(mas[0]) - 4)}
    schedule_classes = {name: {day: {i: None for i in range(1, 9)} for day in days} for name in info_about_classes}
    # Учителя
    teachers_list = [i[0] for i in mas[1:-1]]
    info_about_teachers = {}
    for name in teachers_list:
        vr = {}
        info_about_subjects = {}
        for row in mas[1:-1]:
            if row[0] == name:
                subject = row[1]
                hour_for_classes = {}
                for i in range(2, len(mas[0]) - 4):
                    if row[i] != "-":
                        hour_for_classes[list(info_about_classes.keys())[i - 2]] = row[i]
                # Исключения для "Сдвоенные уроки"
                if ";" in row[-3]:
                    st = row[-3].split("; ")
                    exceptions_double_lessons = {name.split(", ")[0]: int(name.split(", ")[1]) for name in st}
                elif len(row[-3]) > 1:
                    st = row[-3].split(", ")
                    exceptions_double_lessons = {st[0]: int(st[1])}
                else:
                    exceptions_double_lessons = "-"
                # Исключения для "Уроки по группам"
                if ";" in row[-1]:
                    st = row[-1].split("; ")
                    exceptions_lesson_by_groups = {name.split(", ")[0]: (name.split(", ")[1], name.split(", ")[2], int(name.split(", ")[3])) for name in st}
                elif len(row[-1]) > 1:
                    st = row[-1].split(", ")
                    exceptions_lesson_by_groups = {st[0]: (st[1], st[2], int(st[3]))}
                else:
                    exceptions_lesson_by_groups = "-"
                info_about_subjects[subject] = {"Классы": hour_for_classes, 
                                                "Сдвоенные уроки": {"Значение": True if row[-4].lower() == "да" else False, "Исключения": exceptions_double_lessons}, 
                                                "Уроки по группам": {"Значение": True if row[-2].lower() == "да" else False, "Исключения": exceptions_lesson_by_groups}}
        vr["Предметы"] = info_about_subjects
        info_about_teachers[name] = vr
    schedule_teachers = {name: {day: {i: None for i in range(1, 9)} for day in days} for name in info_about_teachers}
    print(info_about_teachers["Терехова М.Р."])

days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']
get_data_from_billing(days)